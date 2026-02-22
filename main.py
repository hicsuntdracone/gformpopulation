# ─────────────────────────────────────────────
#  main.py  —  Entry point, window layout,
#              menu bar, key bindings
# ─────────────────────────────────────────────

import sys
import os
import tkinter as tk
from tkinter import ttk, filedialog, simpledialog, messagebox
from urllib.parse import urlparse, parse_qs
from openpyxl import load_workbook

import state

# ── Helper ────────────────────────────────────
def resource_path(relative_path):
    base = getattr(sys, "_MEIPASS", os.path.abspath("."))
    return os.path.join(base, relative_path)

# ── Root window ───────────────────────────────
root = tk.Tk()
root.title("Google Form URL Generator")
root.geometry("1280x720")

try:
    root.iconbitmap(resource_path("logo.ico"))
except Exception:
    pass

# ── Inject root + StringVars into state ───────
state.root         = root
state.excel_path_var = tk.StringVar()
state.url_var        = tk.StringVar()
state.progress_var   = tk.DoubleVar()

# ── Late imports ───
from core import (generate_urls_threaded, start_bulk_submission,
                  stop_submission, undo, redo)
from fileops import (save_project, save_project_as, load_project,
                     export_as_txt, export_as_xlsx, export_as_csv)
from tools.randomizer import randomizer_gui
from tools.mapper     import map_likert_gui
from tools.histogram  import histogram_gui
from ui.spreadsheet   import update_spreadsheet_view
from ui.mapping       import rebuild_mapping_ui, _bind_scroll, _unbind_scroll, _on_mousewheel


# ─────────────────────────────────────────────
#  Load & Prepare
# ─────────────────────────────────────────────

def handle_load_button():
    global workbook

    # --- Parse URL ---
    try:
        state.parsed_url = urlparse(state.url_var.get())
        state.query_dict = parse_qs(state.parsed_url.query)
        state.entry_ids  = [k for k in state.query_dict if k.startswith("entry.")]
        if not state.entry_ids:
            messagebox.showerror("Error", "URL tidak mengandung Entry ID valid!")
            return
    except Exception:
        messagebox.showerror("Error", "URL tidak valid!")
        return

    # --- Load Excel ---
    try:
        workbook    = load_workbook(state.excel_path_var.get(), data_only=True)
        state.ws    = workbook.active
        if state.ws is None:
            messagebox.showerror("Error", "Gagal membaca worksheet!")
            return

        first_row = list(state.ws.iter_rows(min_row=1, max_row=1, values_only=True))
        if not first_row or not first_row[0]:
            messagebox.showerror("Error", "Excel kosong!")
            return

        state.excel_headers = [str(c.value).strip() for c in state.ws[1] if c.value is not None]
        state.columns       = state.excel_headers + ["NULL"]

        state.excel_data = [
            {col: val for col, val in zip(state.excel_headers, row)}
            for row in state.ws.iter_rows(min_row=2, values_only=True)
            if any(row)
        ]

        if not state.excel_data:
            num_rows = None
            while num_rows is None:
                val = simpledialog.askstring(
                    "Data Kosong",
                    "Excel hanya berisi header (tidak ada baris data).\n"
                    "Masukkan jumlah baris responden yang ingin di-generate:"
                )
                if val is None:
                    return
                if val.isdigit() and int(val) >= 1:
                    num_rows = int(val)
                else:
                    messagebox.showerror("Error", "Masukkan angka >= 1!")
            state.excel_data = [{col: None for col in state.excel_headers}
                                for _ in range(num_rows)]

        state.mapping_vars.clear()
        for idx, eid in enumerate(state.entry_ids):
            var = tk.StringVar()
            var.set(state.excel_headers[idx] if idx < len(state.excel_headers) else "NULL")
            state.mapping_vars[eid] = var

    except Exception as e:
        messagebox.showerror("Error", f"Excel Error: {e}")
        return

    update_spreadsheet_view()

    # --- Ask for section count ---
    num_sections = None
    while num_sections is None:
        val = simpledialog.askstring("Input", "Masukkan jumlah section di form (angka >=1):")
        if val is None:
            return
        if val.isdigit() and int(val) >= 1:
            num_sections = int(val)
        else:
            messagebox.showerror("Error", "Input tidak valid!")

    state.num_sections = num_sections
    page_history     = ",".join(str(i) for i in range(num_sections))
    state.parsed_url = state.parsed_url._replace(
        path=state.parsed_url.path.replace("viewform", "formResponse")
    )
    state.query_dict.pop("usp", None)
    state.query_dict["pageHistory"] = [page_history]

    rebuild_mapping_ui()


def select_file():
    path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if path:
        state.excel_path_var.set(path)


# ─────────────────────────────────────────────
#  Menu
# ─────────────────────────────────────────────

menubar        = tk.Menu(root)
file_menu      = tk.Menu(menubar, tearoff=0)
export_submenu = tk.Menu(file_menu, tearoff=0)
tools_menu     = tk.Menu(menubar, tearoff=0)
edit_menu      = tk.Menu(menubar, tearoff=0)

export_submenu.add_command(label="Export as Text (.txt)",  command=export_as_txt)
export_submenu.add_command(label="Export as Excel (.xlsx)", command=export_as_xlsx)
export_submenu.add_command(label="Export as CSV (.csv)",   command=export_as_csv)

menubar.add_cascade(label="File", menu=file_menu)
file_menu.add_command(label="New Project",      accelerator="Ctrl+N",       command=lambda: os.startfile(__file__))
file_menu.add_command(label="Open Project",     accelerator="Ctrl+O",       command=load_project)
file_menu.add_command(label="Save Project",     accelerator="Ctrl+S",       command=save_project)
file_menu.add_command(label="Save Project As...", accelerator="Ctrl+Shift+S", command=save_project_as)
file_menu.add_separator()
file_menu.add_cascade(label="Export", menu=export_submenu)
file_menu.add_separator()
file_menu.add_command(label="Exit", command=root.quit)

menubar.add_cascade(label="Edit", menu=edit_menu)
edit_menu.add_command(label="Undo", accelerator="Ctrl+Z", command=undo)
edit_menu.add_command(label="Redo", accelerator="Ctrl+Y", command=redo)

menubar.add_cascade(label="Tools", menu=tools_menu)
tools_menu.add_command(label="Mass Likert Randomizer",          command=randomizer_gui)
tools_menu.add_command(label="Likert Scale Mapper",             command=map_likert_gui)
tools_menu.add_separator()
tools_menu.add_command(label="📊 Data Visualizer (Histogram)",  command=histogram_gui)

root.config(menu=menubar)

# ─────────────────────────────────────────────
#  Key Bindings
# ─────────────────────────────────────────────

root.bind_all("<Control-s>",       lambda e: save_project())
root.bind_all("<Control-o>",       lambda e: load_project())
root.bind_all("<Control-n>",       lambda e: os.startfile(__file__))
root.bind_all("<Control-Shift-s>", lambda e: save_project_as())
root.bind_all("<Control-z>",       undo)
root.bind_all("<Control-y>",       redo)

# ─────────────────────────────────────────────
#  Layout — Top Frame (Configuration)
# ─────────────────────────────────────────────

top_frame = tk.LabelFrame(root, text=" 1. Configuration ", padx=10, pady=10)
top_frame.pack(side="top", fill="x", padx=10, pady=5)

tk.Label(top_frame, text="File Excel:").grid(row=0, column=0, sticky="w")
tk.Entry(top_frame, textvariable=state.excel_path_var, width=80).grid(row=0, column=1, padx=10)
tk.Button(top_frame, text="Browse", command=select_file, width=10).grid(row=0, column=2)

tk.Label(top_frame, text="URL Form:").grid(row=1, column=0, sticky="w", pady=5)
tk.Entry(top_frame, textvariable=state.url_var, width=80).grid(row=1, column=1, padx=10)
tk.Button(top_frame, text="Load & Prepare", command=handle_load_button,
          bg="blue", fg="white", width=12).grid(row=1, column=2)

# ─────────────────────────────────────────────
#  Layout — Main Workspace
# ─────────────────────────────────────────────

main_container = tk.Frame(root)
main_container.pack(side="top", fill="both", expand=True, padx=10)

# Column 0: Field Mapping (fixed width, scrollable)
map_container = tk.Frame(main_container, width=350)
map_container.grid(row=0, column=0, sticky="nsew")
map_container.grid_propagate(False)

map_frame    = tk.LabelFrame(map_container, text=" 2. Field Mapping ")
map_frame.pack(fill="both", expand=True)

map_v_scroll = tk.Scrollbar(map_frame, orient="vertical")
map_h_scroll = tk.Scrollbar(map_frame, orient="horizontal")
canvas       = tk.Canvas(map_frame, highlightthickness=0,
                         yscrollcommand=map_v_scroll.set,
                         xscrollcommand=map_h_scroll.set)
map_v_scroll.config(command=canvas.yview)
map_h_scroll.config(command=canvas.xview)
scrollable_frame = tk.Frame(canvas)

canvas.grid(row=0, column=0, sticky="nsew")
map_v_scroll.grid(row=0, column=1, sticky="ns")
map_h_scroll.grid(row=1, column=0, sticky="ew")
map_frame.grid_rowconfigure(0, weight=1)
map_frame.grid_columnconfigure(0, weight=1)

scrollable_frame.bind("<Configure>",
                      lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")

# Inject canvas + scrollable_frame into state for use by mapping.py
state.canvas          = canvas
state.scrollable_frame = scrollable_frame

canvas.bind("<Enter>", _bind_scroll)
canvas.bind("<Leave>", _unbind_scroll)

# Column 1: Data Preview
sheet_container = tk.LabelFrame(main_container, text=" 3. Data Preview ", padx=5, pady=5)
sheet_container.grid(row=0, column=1, sticky="nsew", padx=5)
state.sheet_container = sheet_container

# Column 2: Execution & Log
right_container = tk.Frame(main_container, width=200)
right_container.grid(row=0, column=2, sticky="nsew")
right_container.grid_propagate(False)

action_frame  = tk.LabelFrame(right_container, text=" 4. Execution ", padx=10, pady=10)
action_frame.pack(fill="x", side="top")

btn_font      = ("Arial", 9)
bold_btn_font = ("Arial", 9, "bold")

tk.Button(action_frame, text="Generate URLs",   command=generate_urls_threaded,
          bg="green",   fg="white", font=btn_font).pack(fill="x", pady=2)
tk.Button(action_frame, text="🚀 START SUBMIT", command=start_bulk_submission,
          bg="#d32f2f", fg="white", font=bold_btn_font).pack(fill="x", pady=10)
tk.Button(action_frame, text="Stop Submission", command=stop_submission,
          bg="gray",    fg="white", font=btn_font).pack(fill="x", pady=2)

log_frame = tk.LabelFrame(right_container, text=" System Log ", padx=5, pady=5)
log_frame.pack(fill="both", expand=True, pady=10)

preview_text = tk.Text(log_frame, font=("Consolas", 8), bg="#f0f0f0", width=1, wrap="none")
preview_text.pack(fill="both", expand=True)
log_h_scroll = tk.Scrollbar(log_frame, orient="horizontal", command=preview_text.xview)
preview_text.configure(xscrollcommand=log_h_scroll.set)
log_h_scroll.pack(side="bottom", fill="x")

state.preview_text = preview_text  # inject into state

main_container.grid_columnconfigure(0, weight=0)
main_container.grid_columnconfigure(1, weight=1)
main_container.grid_columnconfigure(2, weight=0)
main_container.grid_rowconfigure(0, weight=1)

# ─────────────────────────────────────────────
#  Layout — Footer
# ─────────────────────────────────────────────

footer_frame = tk.Frame(root, height=30)
footer_frame.pack(side="bottom", fill="x")

ttk.Progressbar(footer_frame, variable=state.progress_var,
                maximum=100, length=500).pack(pady=5)
tk.Label(footer_frame, text="© 2026 PT Kenikmatan | Status: Ready",
         font=("Arial", 8), fg="gray").pack()

# ─────────────────────────────────────────────
root.mainloop()
